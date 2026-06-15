import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
from datetime import datetime
import os
from pathlib import Path
# ── Load data ──────────────────────────────────────────────────────────────────

bom = pd.read_csv(r'D:\Python_QA_automation\Deval_data_pro\extracted\BOM.csv')
inv=pd.read_csv(r'D:\Python_QA_automation\Deval_data_pro\extracted\Inventory.csv')
mat=pd.read_csv(r'D:\Python_QA_automation\Deval_data_pro\extracted\Purchase_Order.csv',
dtype={
        'Destination_Job_Number': str,
        'Sub_Assembly_Job_Number': str
    })
routing = pd.read_csv(r'D:\Python_QA_automation\Deval_data_pro\extracted\Routing.csv')

# Normalize types
bom['Job_Number'] = bom['Job_Number'].astype(str).str.strip()
bom['Parent_Job_Number'] = bom['Parent_Job_Number'].astype(str).str.strip()
bom['Parent_Job_Number'] = bom['Parent_Job_Number'].replace('nan', np.nan)
bom['Order_Number'] = bom['Order_Number'].astype(str).str.strip()
bom['Part_Number'] = bom['Part_Number'].astype(str).str.strip()
bom['Quantity_Open'] = pd.to_numeric(bom['Quantity_Open'], errors='coerce').fillna(0)

inv['Job_Number'] = inv['Job_Number'].astype(str).str.strip()
inv['Part_Number'] = inv['Part_Number'].astype(str).str.strip()
inv['Outside_Service'] = pd.to_numeric(inv['Outside_Service'], errors='coerce').fillna(1)
inv['Stocking_Quantity'] = pd.to_numeric(inv['Stocking_Quantity'], errors='coerce').fillna(0)

mat['Job_Number'] = mat['Job_Number'].astype(str).str.strip()
mat['Status'] = mat['Status'].astype(str).str.strip()

routing['Job_Number'] = routing['Job_Number'].astype(str).str.strip()
routing['Status'] = routing['Status'].astype(str).str.strip()
routing['Step_Number'] = pd.to_numeric(routing['Step_Number'], errors='coerce')

# ── Build parent-child hierarchy lookup ───────────────────────────────────────
# For each job: find super parent (job with no Parent_Job_Number in same order)
def get_super_parent_map(bom_df):
    """Returns dict: job_number -> super_parent_job_number"""
    # Build adjacency: child -> parent
    child_to_parent = {}
    for _, row in bom_df.iterrows():
        if pd.notna(row['Parent_Job_Number']) and row['Parent_Job_Number'] != '':
            child_to_parent[row['Job_Number']] = row['Parent_Job_Number']

    def find_root(job):
        visited = set()
        cur = job
        while cur in child_to_parent:
            if cur in visited:
                break
            visited.add(cur)
            cur = child_to_parent[cur]
        return cur

    super_parent_map = {}
    for job in bom_df['Job_Number'].unique():
        super_parent_map[job] = find_root(job)
    return super_parent_map, child_to_parent

super_parent_map, child_to_parent = get_super_parent_map(bom)

# For each super parent: get all descendant jobs (direct children + sub-children)
def get_all_descendants(super_parent, child_to_parent):
    """Returns set of all job_numbers where super_parent is the root."""
    # Build parent->children map
    parent_to_children = {}
    for child, parent in child_to_parent.items():
        parent_to_children.setdefault(parent, set()).add(child)

    result = {super_parent}
    queue = [super_parent]
    while queue:
        cur = queue.pop()
        for child in parent_to_children.get(cur, []):
            if child not in result:
                result.add(child)
                queue.append(child)
    return result

# Precompute descendants for all super parents
all_super_parents = set(sp for sp in super_parent_map.values())
sp_descendants = {}
for sp in all_super_parents:
    sp_descendants[sp] = get_all_descendants(sp, child_to_parent)

# ── Precompute: Allocated (inventory) ─────────────────────────────────────────
# Filter: Location_Code not null, Purchase_Order_Number is null, Outside_Service==0, Date_Posted not null
inv_filtered = inv[
    inv['Location_Code'].notna() &
    inv['Purchase_Order_Number'].isna() &
    (inv['Outside_Service'] == 0) &
    inv['Date_Posted'].notna()
].copy()

# Map job->order from BOM
job_to_order = bom.drop_duplicates('Job_Number').set_index('Job_Number')['Order_Number'].to_dict()
job_to_part = bom.drop_duplicates('Job_Number').set_index('Job_Number')['Part_Number'].to_dict()

# For Allocated: for each job in an order, sum stocking_quantity where inv job belongs to same order AND part matches
# Per instructions: check ALL job numbers in the order, match part_number from BOM for that job
inv_filtered = inv_filtered.copy()
inv_filtered['Order_Number'] = inv_filtered['Job_Number'].map(job_to_order)

def calc_allocated(job_number, order_number, part_number):
    # All inventory rows within same order with matching part number
    mask = (
        (inv_filtered['Order_Number'] == str(order_number)) &
        (inv_filtered['Part_Number'] == str(part_number))
    )
    return inv_filtered.loc[mask, 'Stocking_Quantity'].sum()

# ── Precompute: Routing info per job ──────────────────────────────────────────
routing_grouped = routing.groupby('Job_Number')

def get_routing_info(job_number):
    """Returns (max_step_finished, all_finished, finished_count, total_count)"""
    if job_number not in routing_grouped.groups:
        return False, False, 0, 0
    grp = routing_grouped.get_group(job_number)
    statuses = grp['Status'].str.strip().str.lower()
    all_finished = (statuses == 'finished').all()
    max_step = grp.loc[grp['Step_Number'].idxmax(), 'Status'].strip().lower() == 'finished'
    finished_count = (statuses == 'finished').sum()
    total_count = len(grp)
    return max_step, all_finished, int(finished_count), int(total_count)

# ── Precompute: Current routing step per job ──────────────────────────────────
def get_work_center_label(row):
    """Return Work_Center if present, else Vendor_Code, else empty string."""
    wc = str(row.get('Work_Center', '') or '').strip()
    vc = str(row.get('Vendor_Code', '') or '').strip()
    if wc and wc.lower() != 'nan':
        return wc
    if vc and vc.lower() != 'nan':
        return vc
    return ''

def calc_current_routing_step(job_number, routing_pct):
    """Derive current routing step label based on rules 3.1 – 3.7."""
    # 3.4 – job not in routing sheet
    if job_number not in routing_grouped.groups:
        return 'No routing Info'

    grp = routing_grouped.get_group(job_number).copy()
    grp = grp.dropna(subset=['Step_Number'])
    grp = grp.sort_values('Step_Number')
    grp['_status_lower'] = grp['Status'].str.strip().str.lower()

    finished_rows = grp[grp['_status_lower'] == 'finished']
    current_rows  = grp[grp['_status_lower'] == 'current']
    all_finished  = len(finished_rows) == len(grp) and len(grp) > 0

    def fmt(step_row, pct):
        step_num = int(step_row['Step_Number'])
        wc = get_work_center_label(step_row)
        pct_str = f"{pct}%"
        return f"Step {step_num} | {pct_str} | {wc}" if wc else f"Step {step_num} | {pct_str}"

    # 3.1 – routing 100%: all steps finished → show last step
    if routing_pct >= 100 and all_finished:
        last_row = grp.iloc[-1]
        return fmt(last_row, 100)

    # Also handle routing 100% via other conditions (Open Qty=0 etc.) – show last step
    if routing_pct >= 100:
        last_row = grp.iloc[-1]
        return fmt(last_row, 100)

    # 3.3 – routing 0%, no Finished, no Current → show first step
    if routing_pct == 0 and len(finished_rows) == 0 and len(current_rows) == 0:
        first_row = grp.iloc[0]
        return fmt(first_row, 0)

    # 3.5 – routing 0%, no Finished, 2+ Current → use latest Current step
    if routing_pct == 0 and len(finished_rows) == 0 and len(current_rows) >= 1:
        latest_current = current_rows.iloc[-1]
        return fmt(latest_current, 0)

    # 3.6 – routing 0%, no Finished, exactly 1 Current
    if routing_pct == 0 and len(finished_rows) == 0 and len(current_rows) == 1:
        curr_row = current_rows.iloc[0]
        return fmt(curr_row, 0)

    # 3.7 – routing between 0–100%, has Finished steps → find next step after latest Finished
    if 0 < routing_pct < 100 and len(finished_rows) > 0:
        latest_finished_step = finished_rows['Step_Number'].max()
        next_steps = grp[grp['Step_Number'] > latest_finished_step]
        if not next_steps.empty:
            next_row = next_steps.iloc[0]
            return fmt(next_row, routing_pct)
        else:
            # No next step — show last finished step
            last_finished = finished_rows[finished_rows['Step_Number'] == latest_finished_step].iloc[0]
            return fmt(last_finished, routing_pct)

    # Fallback: show first step with current routing pct
    first_row = grp.iloc[0]
    return fmt(first_row, routing_pct)

# ── Precompute: Material counts per job ───────────────────────────────────────
mat_job_total = mat.groupby('Job_Number').size().to_dict()
mat_job_posted = mat[mat['Status'].str.lower() == 'posted'].groupby('Job_Number').size().to_dict()

# ── Main calculation loop ──────────────────────────────────────────────────────
# Get unique jobs ordered by Order_Number, then Job_Number
jobs_df = bom[['Order_Number','Job_Number','Part_Number','Part_Description','Quantity_Open','Parent_Job_Number']].drop_duplicates('Job_Number').copy()
jobs_df = jobs_df.sort_values(['Order_Number','Job_Number']).reset_index(drop=True)

# Routing step numbers (for column headers)
all_steps = sorted(routing['Step_Number'].dropna().unique().astype(int))

# Build routing pivot: job -> {step: status}
routing_pivot = routing.pivot_table(index='Job_Number', columns='Step_Number', values='Status', aggfunc='first')
routing_pivot.columns = [int(c) for c in routing_pivot.columns]

# ── Pass 1: compute Allocated and Open_Qty for all jobs (needed for status logic)
allocated_vals = {}
open_qty_vals = {}
for _, row in jobs_df.iterrows():
    jn = row['Job_Number']
    on = row['Order_Number']
    pn = row['Part_Number']
    alloc = calc_allocated(jn, on, pn)
    allocated_vals[jn] = alloc
    qty_open = row['Quantity_Open']
    open_qty_raw = qty_open - alloc
    open_qty_vals[jn] = max(0, open_qty_raw)

# ── Pass 2: compute Routing%, Routing status, Mat%, Inv.status, Mat.status, Fulfilled
def round_pct(val):
    """Standard rounding: 11.5->12, 92.4->92"""
    return math.floor(val + 0.5)  # standard rounding

def calc_routing_pct(job_number, quantity_open, open_qty):
    """Compute routing %"""
    # Condition 1: force 100%
    if quantity_open == 0 or open_qty == 0:
        return 100
    max_step_finished, all_finished, _, _ = get_routing_info(job_number)
    if max_step_finished or all_finished:
        return 100

    # Determine if super parent
    sp = super_parent_map.get(job_number)
    is_super_parent = (sp == job_number and pd.isna(jobs_df.loc[jobs_df['Job_Number']==job_number, 'Parent_Job_Number'].values[0] if job_number in jobs_df['Job_Number'].values else np.nan))

    if is_super_parent:
        family = sp_descendants.get(job_number, {job_number})
        total = sum(get_routing_info(j)[3] for j in family)
        finished = sum(get_routing_info(j)[2] for j in family)
    else:
        _, _, finished, total = get_routing_info(job_number)

    if total == 0:
        return 0
    return round_pct((finished / total) * 100)

def calc_routing_status(job_number, quantity_open, open_qty, routing_pct):
    """Compute routing status"""
    if quantity_open == 0 or open_qty == 0:
        return 'Completed'
    max_step_finished, all_finished, _, _ = get_routing_info(job_number)
    if max_step_finished or all_finished:
        return 'Completed'
    if routing_pct == 0:
        return 'Not Yet Started'
    if routing_pct >= 100:
        return 'Completed'
    return 'In Progress'

def calc_mat_pct(job_number, quantity_open, open_qty, routing_pct_val):
    """Compute Mat%"""
    if quantity_open == 0 or open_qty == 0 or routing_pct_val >= 100:
        return 100

    sp = super_parent_map.get(job_number)
    parent_val = jobs_df.loc[jobs_df['Job_Number']==job_number, 'Parent_Job_Number'].values
    is_super_parent = (sp == job_number and (len(parent_val) == 0 or pd.isna(parent_val[0])))

    if is_super_parent:
        family = sp_descendants.get(job_number, {job_number})
        total = sum(mat_job_total.get(j, 0) for j in family)
        posted = sum(mat_job_posted.get(j, 0) for j in family)
    else:
        total = mat_job_total.get(job_number, 0)
        posted = mat_job_posted.get(job_number, 0)

    if total == 0:
        return 0
    return round_pct((posted / total) * 100)

def calc_inv_status(open_qty, quantity_open, allocated):
    if open_qty == 0 or quantity_open == 0:
        return 'Completed'
    if quantity_open == 0:
        return 'Completed'
    ratio = (allocated / quantity_open) * 100 if quantity_open > 0 else 0
    if ratio == 0:
        return 'Not Yet Started'
    if ratio >= 100:
        return 'Completed'
    return 'In Progress'

def calc_mat_status(open_qty, quantity_open, routing_status, mat_pct):
    if open_qty == 0 or quantity_open == 0 or routing_status == 'Completed':
        return 'Completed'
    if mat_pct == 0:
        return 'Not Yet Started'
    if mat_pct >= 100:
        return 'Completed'
    return 'In Progress'

# ── Build records ─────────────────────────────────────────────────────────────
records = []
for _, row in jobs_df.iterrows():
    jn = row['Job_Number']
    on = row['Order_Number']
    pn = row['Part_Number']
    pd_desc = row['Part_Description']
    qty_open = row['Quantity_Open']
    alloc = allocated_vals[jn]
    open_qty = open_qty_vals[jn]

    routing_pct = calc_routing_pct(jn, qty_open, open_qty)
    routing_status = calc_routing_status(jn, qty_open, open_qty, routing_pct)
    mat_pct = calc_mat_pct(jn, qty_open, open_qty, routing_pct)
    inv_status = calc_inv_status(open_qty, qty_open, alloc)
    mat_status = calc_mat_status(open_qty, qty_open, routing_status, mat_pct)
    current_routing_step = calc_current_routing_step(jn, routing_pct)

    fulfilled = 'YES' if (inv_status == 'Completed' or routing_status == 'Completed') else 'NO'

    rec = {
        'Order number': on,
        'Job number': jn,
        'Part number': pn,
        'Part description': pd_desc,
        'Inv.status': inv_status,
        'Mat.status': mat_status,
        'Allocated': alloc,
        'Open Qty': open_qty,
        'Fulfilled': fulfilled,
        'Mat%': f"{mat_pct}%",
        'Routing%': f"{routing_pct}%",
        'Routing status': routing_status,
        'Current routing step': current_routing_step,
    }

    # Routing step columns
    if jn in routing_pivot.index:
        job_steps = routing_pivot.loc[jn]
        for step in all_steps:
            rec[f'Step {step}'] = job_steps.get(step, '')
    else:
        for step in all_steps:
            rec[f'Step {step}'] = ''

    records.append(rec)

df_expected = pd.DataFrame(records)

# ── Write to Excel ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Expected'

# Header styling
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_fill = PatternFill('solid', start_color='1F4E79')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Status fills
completed_fill = PatternFill('solid', start_color='C6EFCE')   # green
in_progress_fill = PatternFill('solid', start_color='FFEB9C')  # yellow
not_started_fill = PatternFill('solid', start_color='FFC7CE')  # red
yes_fill = PatternFill('solid', start_color='C6EFCE')
no_fill = PatternFill('solid', start_color='FFC7CE')
finished_fill = PatternFill('solid', start_color='C6EFCE')

columns = list(df_expected.columns)

# Write headers
for col_idx, col_name in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Write data
for row_idx, rec in enumerate(records, 2):
    for col_idx, col_name in enumerate(columns, 1):
        val = rec.get(col_name, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.border = thin_border

        # Conditional formatting
        if col_name in ('Inv.status', 'Mat.status', 'Routing status'):
            if val == 'Completed':
                cell.fill = completed_fill
            elif val == 'In Progress':
                cell.fill = in_progress_fill
            elif val == 'Not Yet Started':
                cell.fill = not_started_fill
            cell.alignment = center_align
        elif col_name == 'Fulfilled':
            if val == 'YES':
                cell.fill = yes_fill
            else:
                cell.fill = no_fill
            cell.alignment = center_align
        elif col_name in ('Mat%', 'Routing%'):
            cell.alignment = center_align
        elif col_name in ('Allocated', 'Open Qty'):
            cell.alignment = center_align
            cell.number_format = '#,##0.##'
        elif col_name.startswith('Step '):
            if str(val).strip().lower() == 'finished':
                cell.fill = finished_fill
            cell.alignment = center_align
            cell.font = Font(name='Arial', size=8)
        else:
            cell.alignment = left_align

# Column widths
col_widths = {
    'Order number': 14, 'Job number': 18, 'Part number': 18,
    'Part description': 28, 'Inv.status': 14, 'Mat.status': 14,
    'Allocated': 10, 'Open Qty': 10, 'Fulfilled': 10,
    'Mat%': 8, 'Routing%': 10, 'Routing status': 14, 'Current routing step': 30,
}
for col_idx, col_name in enumerate(columns, 1):
    col_letter = get_column_letter(col_idx)
    width = col_widths.get(col_name, 9)
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


output_dir = r'D:\Python_QA_automation\Deval_data_pro'
output_path = r'D:\Python_QA_automation\Deval_data_pro\expected_data.csv'

output_path = Path(output_path)  # Convert string to Path object
if output_path.exists():
    output_path.unlink()
    print(f"  Deleted existing file: {output_path.name}")



# Delete any existing expected_data_*.csv files in the folder
for file in os.listdir(output_dir):
    if file.startswith("expected_data_") and file.endswith(".csv"):
        old_file = os.path.join(output_dir, file)
        os.remove(old_file)
        print(f"Deleted old file: {old_file}")

# Save new file with today's date
wb.save(output_path)

print(f"Saved to {output_path}")
print(f"Total jobs: {len(records)}")
print(f"Total columns: {len(columns)}")
#print("\nSample rows:")
#print(df_expected[['Order number','Job number','Inv.status','Mat.status','Allocated','Open Qty','Fulfilled','Mat%','Routing%','Routing status']].head(5).to_string())
