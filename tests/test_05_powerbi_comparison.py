"""
test_powerbi_comparison.py
==========================
Pytest test cases for validating PowerBI grid data against the expected data file.
After all tests run, a full comparison Excel report (comparison_report.xlsx) is
automatically generated in the same directory — identical in structure to the
report produced by rough.py (Comparison Summary, per-column mismatch sheets,
All Columns Detail, Column Stats).



Usage:
    pytest test_powerbi_comparison.py -v
    pytest test_powerbi_comparison.py -v --tb=short
"""

import re
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── File paths  (resolved relative to this test file's directory) ─────────────

XLSX_FILE   = (r"D:\Python_QA_automation\Deval_data_pro\PowerBI_grid_data.xlsx")
CSV_FILE    = (r"D:\Python_QA_automation\Deval_data_pro\expected_data.csv")   # actually an xlsx; loader handles it
REPORT_XLSX = str(r"D:\Python_QA_automation\Deval_data_pro\comparison_report.xlsx")
REPORT_XLSX=Path(REPORT_XLSX)
if REPORT_XLSX.exists():
        REPORT_XLSX.unlink()
        print(f"  Deleted existing file: {REPORT_XLSX.name}")


# ── Column mapping: PowerBI col → Expected col ────────────────────────────────
COLUMN_MAP = {
    "J#":          "Job number",
    "P#":          "Part number",
    "Desc":        "Part description",
    "Inv. Status": "Inv.status",
    "Mat. Status": "Mat.status",
    "Allocated":   "Allocated",
    "Open QTY":    "Open Qty",
    "Fulfilled":   "Fulfilled",
    "Mat.%":       "Mat%",
}

DISPLAY_NAMES = {
    "J#":          "Job Number",
    "P#":          "Part Number",
    "Desc":        "Part Description",
    "Inv. Status": "Inv. Status",
    "Mat. Status": "Mat. Status",
    "Allocated":   "Allocated",
    "Open QTY":    "Open QTY",
    "Fulfilled":   "Fulfilled",
    "Mat.%":       "Mat.%",
}

JOIN_KEY_XLSX = "J#"
JOIN_KEY_CSV  = "Job number"


# ═════════════════════════════════════════════════════════════════════════════
# Helpers
# ═════════════════════════════════════════════════════════════════════════════
def sanitize(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", s)
    return s


def normalize_fulfilled(v: str) -> str:
    return v.strip().upper()


def normalize_mat_pct(v: str) -> str:
    v = v.strip().replace("%", "")
    try:
        f = float(v)
        return str(round(f * 100)) if f <= 1.0 else str(round(f))
    except ValueError:
        return v


def load_powerbi(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str, header=1).fillna("")
    df.columns = [sanitize(c) for c in df.columns]
    for col in df.columns:
        df[col] = df[col].apply(sanitize)
    # Strip PowerBI footer/metadata rows
    job_pattern = re.compile(r"^\d+.*-")
    if JOIN_KEY_XLSX in df.columns:
        df = df[df[JOIN_KEY_XLSX].str.match(job_pattern, na=False)].copy()
    return df


def load_expected(path: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, dtype=str).fillna("")
    except Exception:
        import csv as csv_mod
        with open(path, "r", encoding="latin-1", errors="replace") as f:
            sample = f.read(4096)
        try:
            dialect = csv_mod.Sniffer().sniff(sample, delimiters=",;\t|")
            delim = dialect.delimiter
        except Exception:
            delim = ","
        df = pd.read_csv(path, dtype=str, encoding="latin-1",
                         sep=delim, engine="python", on_bad_lines="skip").fillna("")
    df.columns = [sanitize(c) for c in df.columns]
    for col in df.columns:
        df[col] = df[col].apply(sanitize)
    return df


def _resolve_join_col(df, col):
    for candidate in (col + "_xlsx", col):
        if candidate in df.columns:
            return candidate
    return col


def _assert_no_mismatches(mismatch_df, xlsx_col, csv_col, ax, ac):
    if mismatch_df.empty:
        return
    join_col = _resolve_join_col(mismatch_df, "J#")
    pn_col   = _resolve_join_col(mismatch_df, "P#")
    lines = [
        f"\n  {len(mismatch_df)} mismatch(es) found for '{xlsx_col}' vs '{csv_col}':\n",
        f"  {'#':<5} {'J#':<22} {'P#':<14} {'PowerBI Value':<30} {'Expected Value':<30}",
        "  " + "-" * 100,
    ]
    for i, (idx, row) in enumerate(mismatch_df.iterrows(), 1):
        jn = sanitize(row.get(join_col, ""))
        pn = sanitize(row.get(pn_col, ""))
        xv = ax.loc[idx] if idx in ax.index else ""
        cv = ac.loc[idx] if idx in ac.index else ""
        lines.append(f"  {i:<5} {jn:<22} {pn:<14} {xv:<30} {cv:<30}")
        if i >= 30:
            lines.append(f"  ... ({len(mismatch_df) - 30} more rows truncated)")
            break
    assert False, "\n".join(lines)


# ═════════════════════════════════════════════════════════════════════════════
# Excel report helpers  (ported from rough.py)
# ═════════════════════════════════════════════════════════════════════════════
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_style(cell, font_color="000000", bg=None, bold=False,
                 center=False, size=10, wrap=False):
    cell.font = Font(color=font_color, bold=bold, name="Arial", size=size)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap,
    )
    cell.border = _thin()


def _header_cell(cell, text, bg="2C3E50", fg="FFFFFF", size=10,
                 center=True, wrap=False):
    cell.value = text
    cell.font = Font(color=fg, bold=True, name="Arial", size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap,
    )
    cell.border = _thin()


def _set_col_widths(ws, widths: dict):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _run_compare(df_xlsx, df_csv):
    """Replicate rough.py compare() — returns results dict."""
    merged_outer = df_xlsx.merge(
        df_csv,
        left_on=JOIN_KEY_XLSX, right_on=JOIN_KEY_CSV,
        how="outer", indicator=True,
        suffixes=("_xlsx", "_csv"),
    )
    only_xlsx = merged_outer[merged_outer["_merge"] == "left_only"]
    only_csv  = merged_outer[merged_outer["_merge"] == "right_only"]
    both      = merged_outer[merged_outer["_merge"] == "both"].copy()

    results = {
        "total":       len(both),
        "only_xlsx":   only_xlsx,
        "only_csv":    only_csv,
        "both":        both,
        "col_results": {},
    }

    for xlsx_col, csv_col in COLUMN_MAP.items():
        if xlsx_col == JOIN_KEY_XLSX:
            results["col_results"][xlsx_col] = {
                "display":       DISPLAY_NAMES.get(xlsx_col, xlsx_col),
                "total":         len(both),
                "matched":       len(both),
                "mismatched":    0,
                "match_rate":    1.0,
                "mismatch_rows": pd.DataFrame(),
            }
            continue

        col_x = xlsx_col + "_xlsx" if xlsx_col + "_xlsx" in both.columns else xlsx_col
        col_c = csv_col  + "_csv"  if csv_col  + "_csv"  in both.columns else csv_col

        ax = both[col_x].apply(sanitize) if col_x in both.columns else pd.Series([""] * len(both))
        ac = both[col_c].apply(sanitize) if col_c in both.columns else pd.Series([""] * len(both))

        if xlsx_col == "Fulfilled":
            ax = ax.apply(normalize_fulfilled)
            ac = ac.apply(normalize_fulfilled)
        elif xlsx_col == "Mat.%":
            ax = ax.apply(normalize_mat_pct)
            ac = ac.apply(normalize_mat_pct)

        mask = ax != ac
        mismatch_rows = both[mask].copy()
        mismatch_rows = mismatch_rows.assign(_val_xlsx=ax[mask], _val_csv=ac[mask])

        n_mis   = int(mask.sum())
        n_match = len(both) - n_mis

        results["col_results"][xlsx_col] = {
            "display":       DISPLAY_NAMES.get(xlsx_col, xlsx_col),
            "total":         len(both),
            "matched":       n_match,
            "mismatched":    n_mis,
            "match_rate":    n_match / len(both) if len(both) else 1.0,
            "mismatch_rows": mismatch_rows,
            "col_x":         col_x,
            "col_c":         col_c,
        }

    return results


def build_excel_report(results, out_path: Path):
    """Replicate rough.py build_excel() — 4-sheet styled workbook."""
    if out_path.exists():
        out_path.unlink()

    wb   = Workbook()
    both = results["both"]

    # ── Sheet 1: Comparison Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Comparison Summary"
    ws1.freeze_panes = "A5"
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A1:F1")
    _header_cell(ws1["A1"], "AI Grid vs PowerBI — Column Comparison Summary",
                 bg="1A252F", fg="FFFFFF", size=13)

    ws1.merge_cells("A2:F2")
    src_label = (f"Total Records Compared: {results['total']:,}  |  "
                 f"Source 1: PowerBI xlsx  |  Source 2: Expected file")
    ws1["A2"].value = src_label
    _apply_style(ws1["A2"], bg="2C3E50", font_color="BDC3C7", center=True)
    ws1["A2"].border = _thin()

    ws1.append([])

    col_hdrs = ["Column", "Total Records", "Matching Records",
                "Mismatched Records", "Match %", "Status"]
    ws1.append(col_hdrs)
    for c, h in enumerate(col_hdrs, 1):
        _header_cell(ws1.cell(4, c), h, bg="2980B9")

    for xlsx_col, info in results["col_results"].items():
        n_mis  = info["mismatched"]
        rate   = info["match_rate"]
        ok     = n_mis == 0
        status = "✔ Full Match" if ok else f"⚠ {n_mis} Mismatches"
        row    = [info["display"], info["total"], info["matched"],
                  n_mis, f"{rate*100:.2f}%", status]
        ws1.append(row)
        r = ws1.max_row
        s_bg = "D4EDDA" if ok else "FFF3CD"
        s_fg = "155724" if ok else "856404"
        for c in range(1, 7):
            _apply_style(ws1.cell(r, c), center=(c != 1))
        ws1.cell(r, 6).font  = Font(color=s_fg, bold=True, name="Arial", size=10)
        ws1.cell(r, 6).fill  = PatternFill("solid", start_color=s_bg)
        ws1.cell(r, 6).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(r, 6).border = _thin()

    ws1.append([])
    leg_r = ws1.max_row + 1
    ws1.cell(leg_r, 1).value = "Legend"
    ws1.cell(leg_r, 1).font  = Font(bold=True, name="Arial")
    ws1.cell(leg_r+1, 1).value = "Full Match"
    ws1.cell(leg_r+1, 1).fill = PatternFill("solid", start_color="D4EDDA")
    ws1.cell(leg_r+2, 1).value = "Has Mismatches"
    ws1.cell(leg_r+2, 1).fill = PatternFill("solid", start_color="FFF3CD")

    ws1.append([])
    ws1.append([])
    only_xlsx_jobs = (results["only_xlsx"][JOIN_KEY_XLSX].dropna().tolist()
                      if JOIN_KEY_XLSX in results["only_xlsx"].columns else [])
    only_csv_jobs  = (results["only_csv"][JOIN_KEY_CSV].dropna().tolist()
                      if JOIN_KEY_CSV in results["only_csv"].columns else [])

    r = ws1.max_row + 1
    ws1.merge_cells(f"A{r}:F{r}")
    _header_cell(ws1.cell(r, 1),
                 f"Records only in XLSX ({len(only_xlsx_jobs)})", bg="E67E22")
    for job in only_xlsx_jobs[:50]:
        ws1.append(["  " + sanitize(job)])

    ws1.append([])
    r2 = ws1.max_row + 1
    ws1.merge_cells(f"A{r2}:F{r2}")
    _header_cell(ws1.cell(r2, 1),
                 f"Records only in Expected file ({len(only_csv_jobs)})", bg="8E44AD")
    for job in only_csv_jobs[:50]:
        ws1.append(["  " + sanitize(job)])

    _set_col_widths(ws1, {1:24, 2:16, 3:18, 4:19, 5:12, 6:18})

    # ── Sheet 2: Per-column mismatch detail sheets ────────────────────────────
    for xlsx_col, info in results["col_results"].items():
        if info["mismatched"] == 0:
            continue
        mdf = info["mismatch_rows"]
        if mdf.empty:
            continue

        ws = wb.create_sheet(info["display"][:31])
        ws.freeze_panes = "A5"

        ws.merge_cells("A1:H1")
        _header_cell(ws["A1"],
                     f"{info['display']} — Mismatched Records Detail",
                     bg="1A252F", fg="FFFFFF", size=12)
        ws.row_dimensions[1].height = 24

        ws.merge_cells("A2:H2")
        ws["A2"].value = (f"Mismatched: {info['mismatched']} of {info['total']} records  "
                          f"|  Match Rate: {info['match_rate']*100:.2f}%")
        _apply_style(ws["A2"], bg="2C3E50", font_color="BDC3C7", center=True)
        ws["A2"].border = _thin()
        ws.append([])

        hdrs = ["#", "Job Number", "Part Number", "Part Description",
                f"XLSX Value ({xlsx_col})",
                f"Expected Value ({COLUMN_MAP.get(xlsx_col, '')})",
                "Difference"]
        ws.append(hdrs)
        for c, h in enumerate(hdrs, 1):
            _header_cell(ws.cell(4, c), h, bg="2980B9")

        jn_col   = "J#_xlsx"   if "J#_xlsx"   in mdf.columns else "J#"
        pn_col   = "P#_xlsx"   if "P#_xlsx"   in mdf.columns else "P#"
        desc_col = "Desc_xlsx" if "Desc_xlsx" in mdf.columns else "Desc"

        for seq, (_, row_data) in enumerate(mdf.iterrows(), 1):
            xv   = sanitize(row_data.get("_val_xlsx", ""))
            cv   = sanitize(row_data.get("_val_csv",  ""))
            diff = ""
            if xlsx_col == "Mat.%":
                try:
                    diff = f"{float(xv) - float(cv):.0f}%"
                except Exception:
                    diff = ""
            elif xlsx_col in ("Allocated", "Open QTY"):
                try:
                    diff = str(int(xv) - int(cv))
                except Exception:
                    diff = ""

            ws.append([
                seq,
                sanitize(row_data.get(jn_col,   "")),
                sanitize(row_data.get(pn_col,   "")),
                sanitize(row_data.get(desc_col, "")),
                xv, cv, diff,
            ])
            r = ws.max_row
            for c in range(1, 9):
                _apply_style(ws.cell(r, c), center=(c in [1, 7]))
            ws.cell(r, 5).font = Font(color="C0392B", name="Arial", size=10)
            ws.cell(r, 5).fill = PatternFill("solid", start_color="FADBD8")
            ws.cell(r, 6).font = Font(color="1E8449", name="Arial", size=10)
            ws.cell(r, 6).fill = PatternFill("solid", start_color="D5F5E3")

        _set_col_widths(ws, {1:5, 2:16, 3:14, 4:38, 5:22, 6:22, 7:12, 8:22})

    # ── Sheet 3: All Columns Detail ───────────────────────────────────────────
    ws3 = wb.create_sheet("All Columns Detail")
    ws3.freeze_panes = "A5"

    ws3.merge_cells("A1:K1")
    _header_cell(ws3["A1"], "Full Record Comparison — All Columns",
                 bg="1A252F", fg="FFFFFF", size=12)
    ws3.row_dimensions[1].height = 24

    ws3.merge_cells("A2:K2")
    ws3["A2"].value = (f"All {results['total']:,} matched records shown. "
                       "Green = Match | Red = Mismatch")
    _apply_style(ws3["A2"], bg="2C3E50", font_color="BDC3C7", center=True)
    ws3["A2"].border = _thin()
    ws3.append([])

    col_display_order = list(DISPLAY_NAMES.values()) + ["Mismatch Columns"]
    ws3.append(col_display_order)
    for c, h in enumerate(col_display_order, 1):
        _header_cell(ws3.cell(4, c), h, bg="2980B9")

    # Build per-row mismatch flags
    mismatch_flags = {}
    norm_vals_xlsx = {}

    for xlsx_col, info in results["col_results"].items():
        col_x = info.get("col_x", xlsx_col)
        col_c = info.get("col_c", COLUMN_MAP.get(xlsx_col, xlsx_col))
        ax = (both[col_x].apply(sanitize) if col_x in both.columns
              else pd.Series([""] * len(both)))
        ac = (both[col_c].apply(sanitize) if col_c in both.columns
              else pd.Series([""] * len(both)))

        if xlsx_col == "Fulfilled":
            ax_n = ax.apply(normalize_fulfilled)
            ac_n = ac.apply(normalize_fulfilled)
        elif xlsx_col == "Mat.%":
            ax_n = ax.apply(normalize_mat_pct)
            ac_n = ac.apply(normalize_mat_pct)
        else:
            ax_n, ac_n = ax, ac

        mismatch_flags[xlsx_col]  = ax_n != ac_n
        norm_vals_xlsx[xlsx_col]  = ax_n.reset_index(drop=True)

    both_reset = both.reset_index(drop=True)

    for i in range(len(both_reset)):
        row_data  = both_reset.iloc[i]
        mis_cols  = [DISPLAY_NAMES.get(c, c) for c, flag in mismatch_flags.items()
                     if flag.iloc[i]]
        mis_label = "✔ All Match" if not mis_cols else "⚠ " + ", ".join(mis_cols)
        col_vals  = [norm_vals_xlsx[xc].iloc[i] for xc in DISPLAY_NAMES]

        ws3.append(col_vals + [mis_label])
        r = ws3.max_row
        for c, xlsx_col in enumerate(DISPLAY_NAMES, 1):
            cell     = ws3.cell(r, c)
            cell.value = sanitize(col_vals[c - 1])
            is_match = not mismatch_flags[xlsx_col].iloc[i]
            if is_match:
                _apply_style(cell, bg="EAFAF1")
            else:
                cell.font      = Font(color="922B21", name="Arial", size=10)
                cell.fill      = PatternFill("solid", start_color="FADBD8")
                cell.alignment = Alignment(vertical="center")
                cell.border    = _thin()

        last = ws3.cell(r, len(col_display_order))
        last.value = mis_label
        if not mis_cols:
            _apply_style(last, font_color="1E8449", bg="EAFAF1", center=True)
        else:
            _apply_style(last, font_color="922B21", bg="FADBD8", center=True)

    _set_col_widths(ws3, {1:16, 2:13, 3:34, 4:14, 5:13, 6:11,
                           7:10, 8:11, 9:8,  10:24})

    # ── Sheet 4: Column Stats ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Column Stats")
    ws4.freeze_panes = "A4"

    ws4.merge_cells("A1:E1")
    _header_cell(ws4["A1"], "Per-Column Statistics",
                 bg="1A252F", fg="FFFFFF", size=12)
    ws4.row_dimensions[1].height = 22
    ws4.append([])

    stat_hdrs = ["Column", "Total Line Items", "Matching",
                 "Mismatched", "Match Rate"]
    ws4.append(stat_hdrs)
    for c, h in enumerate(stat_hdrs, 1):
        _header_cell(ws4.cell(3, c), h, bg="2980B9")

    for xlsx_col, info in results["col_results"].items():
        ok   = info["mismatched"] == 0
        rate = f"{info['match_rate']*100:.2f}%"
        ws4.append([info["display"], info["total"], info["matched"],
                    info["mismatched"], rate])
        r = ws4.max_row
        for c in range(1, 6):
            _apply_style(ws4.cell(r, c), center=(c != 1))
        bg = "D4EDDA" if ok else "FFF3CD"
        fg = "155724" if ok else "856404"
        ws4.cell(r, 5).font = Font(color=fg, bold=True, name="Arial", size=10)
        ws4.cell(r, 5).fill = PatternFill("solid", start_color=bg)

    _set_col_widths(ws4, {1:22, 2:16, 3:12, 4:14, 5:12})

    wb.save(out_path)
    print(f"\n  ✅ Comparison report saved → {out_path.resolve()}")


# ═════════════════════════════════════════════════════════════════════════════
# Session-scoped fixtures
# ═════════════════════════════════════════════════════════════════════════════
@pytest.fixture(scope="session")
def df_powerbi():
    return load_powerbi(XLSX_FILE)


@pytest.fixture(scope="session")
def df_expected():
    return load_expected(CSV_FILE)


@pytest.fixture(scope="session")
def merged(df_powerbi, df_expected):
    return df_powerbi.merge(
        df_expected,
        left_on=JOIN_KEY_XLSX,
        right_on=JOIN_KEY_CSV,
        how="inner",
        suffixes=("_xlsx", "_csv"),
    )


@pytest.fixture(scope="session")
def comparison_results(df_powerbi, df_expected):
    """Full outer-join comparison results used by both tests and the report."""
    return _run_compare(df_powerbi, df_expected)


# ── Session-end hook: generate the Excel report once all tests are done ───────
def pytest_sessionfinish(session, exitstatus):
    """
    Hook called by pytest after all tests finish.
    Loads data fresh (fixtures aren't available here) and writes the report.
    """
    try:
        print("\n\n📊 Generating comparison Excel report...")
        df_pb  = load_powerbi(XLSX_FILE)
        df_exp = load_expected(CSV_FILE)
        results = _run_compare(df_pb, df_exp)
        build_excel_report(results, REPORT_XLSX)
    except Exception as exc:
        print(f"\n  ⚠️  Could not generate report: {exc}")


# ═════════════════════════════════════════════════════════════════════════════
# TC-01  Unique J# count matches unique Job number count
# ═════════════════════════════════════════════════════════════════════════════
class TestUniqueJobCount:
    def test_unique_j_number_equals_unique_job_number(self, df_powerbi, df_expected):
        """TC-01: count(unique J#) in PowerBI == count(unique Job number) in expected."""
        unique_j   = df_powerbi[JOIN_KEY_XLSX].nunique()
        unique_job = df_expected[JOIN_KEY_CSV].nunique()
        print(f"\n  Unique J# (PowerBI)        : {unique_j}")
        print(f"  Unique Job number (Expected): {unique_job}")
        assert unique_j == unique_job, (
            f"Unique J# count ({unique_j}) != unique Job number count ({unique_job})"
        )


# ═════════════════════════════════════════════════════════════════════════════
# TC-02  P# matches Part number
# ═════════════════════════════════════════════════════════════════════════════
class TestPartNumber:
    def test_p_number_matches_part_number(self, merged):
        """TC-02: P# (PowerBI) == Part number (Expected) for every matched row."""
        col_x = "P#_xlsx" if "P#_xlsx" in merged.columns else "P#"
        col_c = "Part number_csv" if "Part number_csv" in merged.columns else "Part number"
        ax = merged[col_x].apply(sanitize)
        ac = merged[col_c].apply(sanitize)
        _assert_no_mismatches(merged[ax != ac], "P#", "Part number", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-03  Desc matches Part description
# ═════════════════════════════════════════════════════════════════════════════
class TestDescription:
    def test_desc_matches_part_description(self, merged):
        """TC-03: Desc (PowerBI) == Part description (Expected) for every matched row."""
        col_x = "Desc_xlsx" if "Desc_xlsx" in merged.columns else "Desc"
        col_c = "Part description_csv" if "Part description_csv" in merged.columns else "Part description"
        ax = merged[col_x].apply(sanitize)
        ac = merged[col_c].apply(sanitize)
        _assert_no_mismatches(merged[ax != ac], "Desc", "Part description", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-04  Inv. Status matches Inv.status
# ═════════════════════════════════════════════════════════════════════════════
class TestInvStatus:
    def test_inv_status_matches(self, merged):
        """TC-04: Inv. Status (PowerBI) == Inv.status (Expected) for every matched row."""
        col_x = "Inv. Status_xlsx" if "Inv. Status_xlsx" in merged.columns else "Inv. Status"
        col_c = "Inv.status_csv" if "Inv.status_csv" in merged.columns else "Inv.status"
        ax = merged[col_x].apply(sanitize)
        ac = merged[col_c].apply(sanitize)
        _assert_no_mismatches(merged[ax != ac], "Inv. Status", "Inv.status", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-05  Mat. Status matches Mat.status
# ═════════════════════════════════════════════════════════════════════════════
class TestMatStatus:
    def test_mat_status_matches(self, merged):
        """TC-05: Mat. Status (PowerBI) == Mat.status (Expected) for every matched row."""
        col_x = "Mat. Status_xlsx" if "Mat. Status_xlsx" in merged.columns else "Mat. Status"
        col_c = "Mat.status_csv" if "Mat.status_csv" in merged.columns else "Mat.status"
        ax = merged[col_x].apply(sanitize)
        ac = merged[col_c].apply(sanitize)
        _assert_no_mismatches(merged[ax != ac], "Mat. Status", "Mat.status", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-06  Allocated matches Allocated
# ═════════════════════════════════════════════════════════════════════════════
class TestAllocated:
    def test_allocated_matches(self, merged):
        """TC-06: Allocated (PowerBI) == Allocated (Expected) for every matched row."""
        col_x = "Allocated_xlsx" if "Allocated_xlsx" in merged.columns else "Allocated"
        col_c = "Allocated_csv"  if "Allocated_csv"  in merged.columns else "Allocated"
        ax = merged[col_x].apply(sanitize)
        ac = merged[col_c].apply(sanitize)
        _assert_no_mismatches(merged[ax != ac], "Allocated", "Allocated", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-07  Open QTY matches Open Qty
# ═════════════════════════════════════════════════════════════════════════════
class TestOpenQty:
    def test_open_qty_matches(self, merged):
        """TC-07: Open QTY (PowerBI) == Open Qty (Expected) for every matched row."""
        col_x = "Open QTY_xlsx" if "Open QTY_xlsx" in merged.columns else "Open QTY"
        col_c = "Open Qty_csv"  if "Open Qty_csv"  in merged.columns else "Open Qty"
        ax = merged[col_x].apply(sanitize)
        ac = merged[col_c].apply(sanitize)
        _assert_no_mismatches(merged[ax != ac], "Open QTY", "Open Qty", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-08  Fulfilled matches Fulfilled (case-insensitive YES/NO)
# ═════════════════════════════════════════════════════════════════════════════
class TestFulfilled:
    def test_fulfilled_matches(self, merged):
        """TC-08: Fulfilled (PowerBI) == Fulfilled (Expected), normalised to YES/NO."""
        col_x = "Fulfilled_xlsx" if "Fulfilled_xlsx" in merged.columns else "Fulfilled"
        col_c = "Fulfilled_csv"  if "Fulfilled_csv"  in merged.columns else "Fulfilled"
        ax = merged[col_x].apply(sanitize).apply(normalize_fulfilled)
        ac = merged[col_c].apply(sanitize).apply(normalize_fulfilled)
        _assert_no_mismatches(merged[ax != ac], "Fulfilled", "Fulfilled", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-09  Mat.% matches Mat% (decimal vs percentage normalised)
# ═════════════════════════════════════════════════════════════════════════════
class TestMatPercent:
    def test_mat_pct_matches(self, merged):
        """TC-09: Mat.% (PowerBI decimal 0.875) == Mat% (Expected '88%'), int-normalised."""
        col_x = "Mat.%_xlsx" if "Mat.%_xlsx" in merged.columns else "Mat.%"
        col_c = "Mat%_csv"   if "Mat%_csv"   in merged.columns else "Mat%"
        ax = merged[col_x].apply(sanitize).apply(normalize_mat_pct)
        ac = merged[col_c].apply(sanitize).apply(normalize_mat_pct)
        _assert_no_mismatches(merged[ax != ac], "Mat.%", "Mat%", ax, ac)


# ═════════════════════════════════════════════════════════════════════════════
# TC-10  No records missing from either side
# ═════════════════════════════════════════════════════════════════════════════
class TestRecordCoverage:
    def test_no_records_only_in_powerbi(self, df_powerbi, df_expected):
        """TC-10a: Every J# in PowerBI has a matching Job number in Expected."""
        powerbi_jobs  = set(df_powerbi[JOIN_KEY_XLSX].apply(sanitize).unique())
        expected_jobs = set(df_expected[JOIN_KEY_CSV].apply(sanitize).unique())
        only_in_powerbi = powerbi_jobs - expected_jobs
        assert not only_in_powerbi, (
            f"{len(only_in_powerbi)} J# found only in PowerBI (not in Expected):\n"
            + "\n".join(sorted(only_in_powerbi)[:20])
            + ("\n  ... (truncated)" if len(only_in_powerbi) > 20 else "")
        )

    def test_no_records_only_in_expected(self, df_powerbi, df_expected):
        """TC-10b: Every Job number in Expected has a matching J# in PowerBI."""
        powerbi_jobs  = set(df_powerbi[JOIN_KEY_XLSX].apply(sanitize).unique())
        expected_jobs = set(df_expected[JOIN_KEY_CSV].apply(sanitize).unique())
        only_in_expected = expected_jobs - powerbi_jobs
        assert not only_in_expected, (
            f"{len(only_in_expected)} Job numbers found only in Expected (not in PowerBI):\n"
            + "\n".join(sorted(only_in_expected)[:20])
            + ("\n  ... (truncated)" if len(only_in_expected) > 20 else "")
        )